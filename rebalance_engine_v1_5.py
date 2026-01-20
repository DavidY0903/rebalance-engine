# ===============================================================
# 📘 Rebalance Engine v1.5 — Final Full Fixed Edition v3
# Version: 1.5-final-full-v3 (2025-10-10)
# ---------------------------------------------------------------
# ✅ Restored pointer-based file picker (manual selection)
# ✅ Dynamic output naming (<input>_output.xlsx)
# ✅ Robust I/O + section parsing (from 9-27 baseline)
# ✅ Bilingual Excel sheets and formatting
# ✅ Correct handling of zero / negative cash
# ===============================================================

import os, re, time
import pandas as pd
import numpy as np
from datetime import datetime
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

import json
from pathlib import Path
import requests
from dotenv import load_dotenv
from typing import Optional

MAX_FALLBACK_ENTRIES = 500

load_dotenv(dotenv_path=".env", override=True)
POLYGON_API_KEY = os.getenv("POLYGON_API_KEY")

ALPHA_VANTAGE_KEY = os.getenv("ALPHA_VANTAGE_KEY")

if not ALPHA_VANTAGE_KEY:
    print("❌ ALPHA_VANTAGE_KEY not found. Check .env or environment variables.")

FALLBACK_PATH = Path("price_fallback.json")

if not POLYGON_API_KEY:
    print("❌ POLYGON_API_KEY not found. Check .env or environment variables.")
    
FMP_API_KEY = os.getenv("FMP_API_KEY")

def _fetch_fmp_price(ticker: str) -> Optional[float]:
    if not FMP_API_KEY:
        return None
    try:
        r = requests.get(
            f"https://financialmodelingprep.com/api/v3/quote/{ticker}",
            params={"apikey": FMP_API_KEY},
            timeout=5
        )
        if r.status_code != 200:
            return None

        js = r.json()
        if isinstance(js, list) and js:
            px = js[0].get("price")
            return float(px) if px and px > 0 else None
    except Exception:
        return None

def _load_fallback() -> dict:
    if FALLBACK_PATH.exists():
        try:
            return json.loads(FALLBACK_PATH.read_text())
        except Exception:
            return {}
    return {}

def _save_fallback(data: dict):
    try:
        if len(data) > MAX_FALLBACK_ENTRIES:
            data = dict(list(data.items())[-MAX_FALLBACK_ENTRIES:])
        FALLBACK_PATH.write_text(json.dumps(data, indent=2))
    except Exception:
        pass

def _fetch_polygon_price(ticker: str) -> Optional[float]:
    if not POLYGON_API_KEY:
        return None

    url = f"https://api.polygon.io/v2/aggs/ticker/{ticker}/prev"
    try:
        r = requests.get(
            url,
            params={"adjusted": "true", "apiKey": POLYGON_API_KEY},
            timeout=5
        )
        if r.status_code != 200:
            return None

        js = r.json()
        if js.get("results"):
            px = js["results"][0].get("c")
            return float(px) if px and px > 0 else None
    except Exception:
        return None

    return None

def _fetch_alpha_price(ticker: str) -> Optional[float]:
    """
    Alpha Vantage GLOBAL_QUOTE
    Works well for ETFs (VGT / VIG / XLK / etc.)
    """
    if not ALPHA_VANTAGE_KEY:
        return None

    try:
        r = requests.get(
            "https://www.alphavantage.co/query",
            params={
                "function": "GLOBAL_QUOTE",
                "symbol": ticker,
                "apikey": ALPHA_VANTAGE_KEY
            },
            timeout=8
        )
        if r.status_code != 200:
            return None

        js = r.json()

        if "Note" in js or "Information" in js:
            return None

        quote = js.get("Global Quote", {})
        px = quote.get("05. price")

        if px:
            px = float(px)
            return px if px > 0 else None

    except Exception:
        return None

    return None

def fetch_last_prices(
    tickers: list[str],
    *,
    sleep_between: float = 12.0  # Alpha free tier needs this
) -> dict[str, float]:
    """
    Live-first, HTTP-safe price resolver.
    Polygon → Alpha Vantage → local fallback
    """

    tickers = list(dict.fromkeys(t.strip().upper() for t in tickers if t))
    fallback = _load_fallback()
    resolved: dict[str, float] = {}

    for i, t in enumerate(tickers, start=1):
        print(f"📡 Resolving price {i}/{len(tickers)}: {t}", flush=True)

        # 1 Polygon (primary)
        px = _fetch_polygon_price(t)
        if px is not None:
            print(f"   ✅ Polygon: {px}", flush=True)
            resolved[t] = px
            fallback[t] = px
            _save_fallback(fallback)
            continue

        # 2 Alpha Vantage (secondary, strict rate limit)
        px = _fetch_alpha_price(t)
        time.sleep(sleep_between)

        if px is not None:
            print(f"   🟦 Alpha Vantage: {px}", flush=True)
            resolved[t] = px
            fallback[t] = px
            _save_fallback(fallback)
            continue
        
        # 3 Financial Modeling Prep (cloud-stable)
        px = _fetch_fmp_price(t)
        if px is not None:
            print(f"   🟪 FMP: {px}", flush=True)
            resolved[t] = px
            fallback[t] = px
            _save_fallback(fallback)
            continue

        # 4 Local fallback
        if t in fallback:
            print(f"   🟡 Local fallback: {fallback[t]}", flush=True)
            resolved[t] = float(fallback[t])
            continue

        # 5 Hard fail marker (handled later)
        print(f"   ❌ No price available", flush=True)
        resolved[t] = np.nan

    return resolved

# ---------- Input Handling (Pointer File Picker) ----------
def _pick_input_filename() -> str:
    print("📂 Please select the Excel input file...")
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select the Rebalance Input File",
        filetypes=[("Excel files", "*.xlsx")]
    )
    root.update()
    if not file_path:
        raise FileNotFoundError("❌ No file selected. Please choose an input Excel file.")
    print(f"✅ Selected input file: {os.path.basename(file_path)}")
    return file_path

# ---------- Helpers ----------
def _norm_key(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

def _extract_config(raw_df: pd.DataFrame) -> dict:
    cfg = raw_df.iloc[2:7, [1, 2]].copy()
    cfg.iloc[:, 0] = cfg.iloc[:, 0].astype(str).str.strip().str.lower()
    kv = {_norm_key(k): v for k, v in zip(cfg.iloc[:, 0], cfg.iloc[:, 1])}

    def getv(name, pct=False, yesno=False, default=None, aliases=None):
        keys = [_norm_key(name)]
        if aliases:
            keys += [_norm_key(a) for a in aliases]

        found_key = None
        for k in keys:
            if k in kv:
                found_key = k
                v = kv[k]
                break

        if found_key is None:
            print(f"⚠️ [Config] Key NOT found: {keys}, using default={default}")
            return default

        if yesno:
            result = str(v).strip().lower() in ("yes", "y", "true", "1")
            return result

        if pct:
            result = float(str(v).replace("%", "")) / 100.0
            return result

        try:
            result = float(str(v).replace(",", ""))
            return result
        except Exception:
            return default

    config = {
        "Cash Contribution": getv("cash contribution", default=0.0),
        "Upper Bound": getv("upper bound", pct=True, default=1.0),
        "Lower Bound": getv("lower bound", pct=True, default=0.0),
        "Relax Limit": getv("relax limit", pct=True, default=1.0),
        "Allow Partial Shares": getv(
            "allow partial shares",
            yesno=True,
            default=False,
            aliases=[
                "allow partial shares y/n",
                "allow partial shares yn",
                "允許零股"
            ]
        ),
    }
    return config

def _extract_section(raw_df: pd.DataFrame, title: str, required_cols: list[str]) -> pd.DataFrame:
    m = raw_df[raw_df.iloc[:, 0].astype(str).str.contains(title, na=False)]
    if m.empty:
        raise ValueError(f"❌ Section not found: {title}")
    start = m.index[0]
    header = [str(x).strip() for x in raw_df.iloc[start + 1].tolist()]

    header_map = {}
    for h in header:
        if not h or h.lower() == "nan":
            continue
        if "ticker" in h.lower():
            header_map[h] = "Ticker"
        elif "weight" in h.lower():
            header_map[h] = "Weight"
        elif "quantity" in h.lower() or "shares" in h.lower():
            header_map[h] = "Quantity"
        else:
            header_map[h] = h

    data = []
    for i in range(start + 2, len(raw_df)):
        row = raw_df.iloc[i].tolist()
        if pd.isna(row[0]) or ("Portfolio" in str(row[0]) and i > start + 2):
            break
        data.append(row[:len(header)])

    df = pd.DataFrame(data, columns=header).dropna(how="all").reset_index(drop=True)
    df = df.rename(columns=header_map)
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(f"❌ Missing column '{col}' in section '{title}'. Found header={header}")
    return df

# ---------- Excel Styling ----------
header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
buy_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
sell_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

thin_top = Border(top=Side(style="thin"))
full_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def format_sheet(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="新細明體", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = full_border

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(10, max_len + 2)

def style_action(ws):
    for r in range(2, ws.max_row + 1):
        val = ws[f"B{r}"].value or ws[f"C{r}"].value
        if val in ["Buy", "買進"]:
            for c in ws[r]:
                c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif val in ["Sell", "賣出", "Sell All"]:
            for c in ws[r]:
                c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ---------- Main ----------
def main():
    print("📘 Rebalance Engine v1.5 — Final Full Fixed Edition v3")
    input_path = _pick_input_filename()
    xl = pd.ExcelFile(input_path)
    sheet_name = next((s for s in xl.sheet_names if s.strip().lower() == "current profile"), xl.sheet_names[0])
    raw_input_df = pd.read_excel(input_path, sheet_name=sheet_name, header=None)

    config = _extract_config(raw_input_df)
    target_df = _extract_section(raw_input_df, "Target Portfolio & Weight", ["Ticker", "Weight"])[["Ticker", "Weight"]]
    port_df = _extract_section(raw_input_df, "Current Portfolio & Quantity", ["Ticker", "Quantity"])[["Ticker", "Quantity"]]

    target_df.columns = ["Ticker", "Target Weight (%)"]
    target_df["Target Weight (%)"] = (
        target_df["Target Weight (%)"].astype(str).str.replace("%", "", regex=False).str.replace(",", "", regex=False)
    )
    target_df["Target Weight (%)"] = pd.to_numeric(target_df["Target Weight (%)"], errors="coerce").fillna(0.0)
    if target_df["Target Weight (%)"].sum() > 1.5:
        target_df["Target Weight (%)"] /= 100.0

    port_df.columns = ["Ticker", "Shares"]
    port_df["Shares"] = pd.to_numeric(
        port_df["Shares"].astype(str).str.replace(",", "", regex=False), errors="coerce"
    ).fillna(0.0)

    df = pd.merge(target_df, port_df, on="Ticker", how="outer").fillna({"Target Weight (%)": 0.0, "Shares": 0.0})
    
    price_map = fetch_last_prices(df["Ticker"].tolist())

    df["Last"] = df["Ticker"].map(price_map)
    df["Ask"] = df["Last"]
    df["Bid"] = df["Last"]
    
    missing = df[df["Last"].isna()]["Ticker"].tolist()
    if missing:
        raise ValueError(
            f"❌ Price fetch failed for: {missing}\n"
            f"Polygon + Alpha Vantage + local fallback all failed."
        )

    df["Used Price"] = df["Last"]
    df["Market Value"] = df["Shares"] * df["Used Price"]
    total_value = df["Market Value"].sum() + config["Cash Contribution"]
    df["Target Value"] = df["Target Weight (%)"] * total_value

    # ---------- Allocation math (supports Allow Partial Shares + cash + / 0 / -) ----------
    allow_partial = bool(config.get("Allow Partial Shares", False))
    cash_contrib = float(config.get("Cash Contribution", 0.0))
    
    if cash_contrib < 0:
        print("🔄 Rebalance mode: WITHDRAWAL (negative cash)")
    elif cash_contrib == 0:
        print("🔄 Rebalance mode: PURE REBALANCE (cash-neutral)")
    else:
        print("🔄 Rebalance mode: CONTRIBUTION (positive cash)")
    
    pure_rebalance = (cash_contrib == 0)

    df["Diff Value"] = df["Target Value"] - df["Market Value"]
    
    if cash_contrib < 0:
        df["Target Value"] = np.nan
        df["Diff Value"] = -df["Market Value"]

    # Base rebalance (float first)
    df["Shares to Buy/Sell"] = df["Diff Value"] / df["Used Price"]

    # ---------- Withdrawal mode (cash < 0) ----------
    if cash_contrib < 0:
        needed = abs(cash_contrib)

        df["SellableValue"] = df["Market Value"].clip(lower=0)
        total_sellable = float(df["SellableValue"].sum())

        if total_sellable > 0:
            df["ProportionalToSell"] = df["SellableValue"] / total_sellable
            df["ValueToSell"] = df["ProportionalToSell"] * needed
            df["Shares to Buy/Sell"] = -(df["ValueToSell"] / df["Used Price"])
        else:
            df["Shares to Buy/Sell"] = 0.0

    # Clean tiny noise
    df.loc[df["Shares to Buy/Sell"].abs() < 1e-6, "Shares to Buy/Sell"] = 0.0
    
    PARTIAL_DECIMALS = 2
    FACTOR = 10 ** PARTIAL_DECIMALS

    # ---------- Enforce share granularity ----------
    if not allow_partial:
        df["Shares to Buy/Sell"] = np.where(
            df["Shares to Buy/Sell"] > 0,
            np.floor(df["Shares to Buy/Sell"]),
            np.ceil(df["Shares to Buy/Sell"])
        )
    else:
        # cap first (safety)
        df["Shares to Buy/Sell"] = np.where(
            df["Shares to Buy/Sell"] < -df["Shares"],
            -df["Shares"],
            df["Shares to Buy/Sell"]
        )

        # broker-safe rounding: buys floor, sells ceil (toward zero)
        df["Shares to Buy/Sell"] = np.where(
            df["Shares to Buy/Sell"] > 0,
            np.floor(df["Shares to Buy/Sell"] * FACTOR) / FACTOR,
            np.ceil(df["Shares to Buy/Sell"] * FACTOR) / FACTOR
        )

        # cap again after rounding (just in case rounding crossed boundary)
        df["Shares to Buy/Sell"] = np.where(
            df["Shares to Buy/Sell"] < -df["Shares"],
            -df["Shares"],
            df["Shares to Buy/Sell"]
        )
    
    # ---------- Integer buy allocator (cash-positive, no partial shares) ----------
    if cash_contrib > 0 and not allow_partial:

        remaining_cash = cash_contrib

        # Work on a local view sorted by underweight severity
        while True:
            buy_candidates = df[
                (df["Diff Value"] > 0) &
                (df["Used Price"] <= remaining_cash)
            ].sort_values("Diff Value", ascending=False)

            if buy_candidates.empty:
                break

            idx = buy_candidates.index[0]
            price = df.at[idx, "Used Price"]

            # Buy exactly ONE whole share
            df.at[idx, "Shares to Buy/Sell"] += 1
            remaining_cash -= price

            # Update Diff Value so ranking stays correct
            df.at[idx, "Diff Value"] -= price

            # Safety exit
            min_price = df.loc[df["Used Price"] > 0, "Used Price"].min()
            if remaining_cash < min_price:
                break

    # ---------- Withdrawal top-up (integer shares) ----------
    if cash_contrib < 0 and not allow_partial:
        proceeds = float(
            ((-df["Shares to Buy/Sell"].clip(upper=0)) * df["Used Price"]).sum()
        )
        shortfall = needed - proceeds

        if shortfall > 0:
            sell_candidates = (
                df[(df["Shares"] > 0) & (df["Used Price"] > 0)]
                .sort_values("Market Value", ascending=False)
            )

            for idx in sell_candidates.index:
                if shortfall <= 0:
                    break

                already_selling = int(abs(min(0, df.at[idx, "Shares to Buy/Sell"])))
                remaining = int(df.at[idx, "Shares"]) - already_selling
                if remaining <= 0:
                    continue

                px = float(df.at[idx, "Used Price"])
                extra = int(np.ceil(shortfall / px))
                extra = min(extra, remaining)

                if extra > 0:
                    df.at[idx, "Shares to Buy/Sell"] -= extra
                    shortfall -= extra * px
                    
    # ---------- Pure rebalance cash neutralization ----------
    if pure_rebalance:
        buy_value = (
            df.loc[df["Shares to Buy/Sell"] > 0, "Shares to Buy/Sell"]
            * df["Used Price"]
        ).sum()

        sell_value = (
            -df.loc[df["Shares to Buy/Sell"] < 0, "Shares to Buy/Sell"]
            * df["Used Price"]
        ).sum()

        imbalance = round(buy_value - sell_value, 2)

        if abs(imbalance) > 0.01:
            # Need more sells if imbalance > 0 (buy > sell)
            # Need more buys if imbalance < 0
            if imbalance > 0:
                candidates = df[
                    (df["Shares"] > 0) & (df["Shares to Buy/Sell"] <= 0)
                ].sort_values("Diff Value")
            else:
                candidates = df[
                    df["Shares to Buy/Sell"] >= 0
                ].sort_values("Diff Value", ascending=False)

            if not candidates.empty:
                idx = candidates.index[0]
                px = df.at[idx, "Used Price"]

                delta_shares = imbalance / px

                if allow_partial:
                    delta_shares = (
                        np.floor(delta_shares * FACTOR) / FACTOR
                        if delta_shares > 0
                        else np.ceil(delta_shares * FACTOR) / FACTOR
                    )
                else:
                    delta_shares = (
                        int(np.floor(delta_shares))
                        if delta_shares > 0
                        else int(np.ceil(delta_shares))
                    )

                df.at[idx, "Shares to Buy/Sell"] -= delta_shares
                df["Shares to Buy/Sell"] = np.where(
                    df["Shares to Buy/Sell"] < -df["Shares"],
                    -df["Shares"],
                    df["Shares to Buy/Sell"]
                )

    # ---------- Actions ----------
    df["Action"] = np.where(
        df["Shares to Buy/Sell"] > 0, "Buy",
        np.where(df["Shares to Buy/Sell"] < 0, "Sell", "Hold")
    )

    df.loc[(df["Shares"] == 0) & (df["Target Weight (%)"] > 0), "Action"] = "Buy"
    df.loc[(df["Target Weight (%)"] == 0) & (df["Shares"] > 0), "Action"] = "Sell All"
    df.loc[df["Action"] == "Sell All", "Shares to Buy/Sell"] = -df["Shares"]

    df["Order Type"] = "Limit"
    df["Limit Price"] = np.where(df["Action"].isin(["Buy", "買進"]), df["Ask"], df["Bid"])
    df["Limit Price"] = df["Limit Price"].fillna(df["Used Price"])

    # ---------- Summary & Output ----------
    buy_cash = (df.loc[df["Action"] == "Buy", "Shares to Buy/Sell"] * df["Used Price"]).sum()
    sell_cash = (-df.loc[df["Action"].isin(["Sell", "Sell All"]), "Shares to Buy/Sell"] * df["Used Price"]).sum()
    cash_remaining = round(sell_cash - buy_cash + config["Cash Contribution"], 2)

    # ----- Updated Summary Block (Bilingual + Styled + Wrap Text) -----
    summary = pd.DataFrame({
    "Summary": [
        "買進支付金額 (Cash Used for Buys)",
        "賣出收入 (Cash Received from Sells)",
        "現金結餘 (Net Cash Result)"
    ],
    "Value": [
        round(buy_cash, 2),
        round(sell_cash, 2),
        cash_remaining
    ]
})

    # ----- Create Rebalance Recommendation (Format B) -----
    cols = ["Ticker", "Target Weight (%)", "Shares", "Used Price",
            "Market Value", "Target Value", "Diff Value",
            "Shares to Buy/Sell", "Action", "Order Type"]
    rebalance_df = df[cols].copy()

    total_row = pd.DataFrame({
        "Ticker": ["Total"],
        "Market Value": [df["Market Value"].sum()],
        "Target Value": [df["Target Value"].sum()],
        "Diff Value": [config["Cash Contribution"]]
    })
    rebalance_df = pd.concat([rebalance_df, total_row], ignore_index=True)

    # ----- Excel Output -----
    input_dir = os.path.dirname(input_path) or os.getcwd()
    input_name = os.path.splitext(os.path.basename(input_path))[0]

    m = re.findall(r"\(.*?\)", input_name)
    tag = "".join(m) if m else ""
    timestamp = datetime.now().strftime("%Y-%m-%d %H%M%S")

    output_name = os.path.join(
        input_dir, f"rebalance recommendation {tag} ({timestamp}).xlsx"
    )

    wb = Workbook()
    wb.remove(wb.active)

    sort_order = {"Sell": 0, "Sell All": 0, "Buy": 1, "Hold": 2}
    df_sorted = df.copy()
    df_sorted["sort_key"] = df_sorted["Action"].map(sort_order)
    df_sorted = df_sorted.sort_values(by=["sort_key", "Diff Value"], ascending=[True, False])
    df_sorted = df_sorted.drop(columns=["sort_key"])

    # ----- Action sheet -----
    df_action = df_sorted[["Ticker", "Action", "Shares to Buy/Sell", "Order Type", "Limit Price"]]
    ws_action = wb.create_sheet("action")
    for row in dataframe_to_rows(df_action, index=False, header=True):
        ws_action.append(row)
    format_sheet(ws_action)
    style_action(ws_action)

    # ----- 中文執行 sheet -----
    df_exec = df_action.copy()
    df_exec.columns = ["股票代碼", "動作", "買賣股數", "委託類型", "委託價格"]

    df_exec["動作"] = df_exec["動作"].replace({
        "Buy": "買進",
        "Sell": "賣出",
        "Sell All": "全部賣出",
        "Hold": "持有"
    })
    df_exec["委託類型"] = df_exec["委託類型"].replace({
        "Limit": "限價",
        "Market": "市價"
    })

    ws_exec = wb.create_sheet("執行")
    for row in dataframe_to_rows(df_exec, index=False, header=True):
        ws_exec.append(row)
    format_sheet(ws_exec)
    style_action(ws_exec)

    # ----- Rebalance Recommendation sheet -----
    ordered_rebal = (
        rebalance_df[rebalance_df["Ticker"] != "Total"]
            .set_index("Ticker")
            .reindex(df_action["Ticker"])
            .reset_index()
    )

    total_rebal = rebalance_df[rebalance_df["Ticker"] == "Total"]

    ws_rebal = wb.create_sheet("Rebalance Recommendation")
    for row in dataframe_to_rows(pd.concat([ordered_rebal, total_rebal], ignore_index=True),
                                index=False, header=True):
        ws_rebal.append(row)

    format_sheet(ws_rebal)

    for r in range(2, ws_rebal.max_row + 1):
        ticker_val = str(ws_rebal[f"A{r}"].value or "").strip()
        if ticker_val.lower() == "total":
            for cell in ws_rebal[r]:
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border(top=Side(style="medium"))
            continue

        action_val = ws_rebal[f"I{r}"].value
        if action_val in ("Buy", "買進"):
            for cell in ws_rebal[r]:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif action_val in ("Sell", "賣出", "Sell All"):
            for cell in ws_rebal[r]:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws_rebal.append([])
    ws_rebal.append([])

    # ----- Styled Summary Block (Bilingual + Auto Wrap + Borders + Align) -----
    for row_idx, row in enumerate(dataframe_to_rows(summary, index=False, header=True), start=1):
        ws_rebal.append(row)

    # Apply styles to the entire summary block
    start_row = ws_rebal.max_row - len(summary)  # first summary row
    end_row = ws_rebal.max_row                   # last summary row

    for r in range(start_row, end_row + 1):
        summary_cell = ws_rebal[f"A{r}"]
        value_cell = ws_rebal[f"B{r}"]

        # Summary column formatting
        summary_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )
        summary_cell.font = Font(name="新細明體", bold=False, size=11)
        summary_cell.border = full_border

        # Value column formatting (right-aligned)
        value_cell.alignment = Alignment(
            horizontal="right",
            vertical="center"
        )
        value_cell.font = Font(name="新細明體", bold=True, size=11)
        value_cell.border = full_border

    # ----- Keep original input file content as last sheet -----
    if "current input file" in wb.sheetnames:
        del wb["current input file"]

    ws_input = wb.create_sheet("current input file")

    for row in dataframe_to_rows(raw_input_df, index=False, header=False):
        ws_input.append(row)

    format_sheet(ws_input)

    wb.save(output_name)
    print(f"✅ Output saved: {output_name}")

if __name__ == "__main__":
    main()
